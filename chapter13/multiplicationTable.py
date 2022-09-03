#! python3
# Create a program multiplicationTable.py that takes a number
# N from the command line and creates
# an N×N multiplication table in an Excel spreadsheet

import sys, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

number = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.freeze_panes = 'B2'
bold_font = Font(bold=True)
for n in range (2,number+2):
    sheet[get_column_letter(n)+'1'] = n-1
    sheet[get_column_letter(n)+'1'].font = bold_font
    sheet['A'+str(n)]= n-1
    sheet['A'+str(n)].font= bold_font
for row in range(2, number+2):
    for column in range(2, number+2):
        sheet[get_column_letter(column)+str(row)]=f'={get_column_letter(column)}1*A{str(row)}'
wb.save('multiplication.xlsx')
