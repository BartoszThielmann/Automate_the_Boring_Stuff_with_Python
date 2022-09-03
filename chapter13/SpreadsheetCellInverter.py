#! python3
# Spreadsheet Cell Inverter
# Write a program to invert the row and column 
# of the cells in the spreadsheet
# usage: in terminal type: python3 SpreadsheetCellInverter.py "filename to invert"
# the file needs to be in the same folder as SpreadsheetCellInverter.py

import openpyxl,sys
load = sys.argv[1]
wb = openpyxl.load_workbook(load)
sheet = wb['Sheet1']

# sheetData=[[ITEM, Eggplant, Cucumber, Green cabbage][SOLD, 334, 252, 238]]
sheet_data = []
for ccolumn in range(1, sheet.max_column + 1):
	sheet_data.append([])
	for crow in range(1, sheet.max_row + 1):
		sheet_data[ccolumn-1].append(sheet.cell(row = crow, column = ccolumn).value)
		
for mrows in range(len(sheet_data)):
	for mcolumns in range(len(sheet_data[mrows])):
		sheet.cell(row = mrows + 1, column = mcolumns + 1).value = sheet_data[mrows][mcolumns]
		
for i in range(len(sheet_data) + 1, sheet.max_row + 1):
	for j in range(1, ccolumn+1):
		sheet.cell(row = i, column = j).value=''

wb.save('inverter_result.xlsx')
