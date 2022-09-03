#! python3
# Create a program blankRowInserter.py that takes two integers
# and a filename string as command line arguments.
# Let’s call the first integer N and the second integer M.
# Starting at row N, the program should insert M blank rows into the spreadsheet.

import openpyxl, sys

filename = sys.argv[1]
N = int(sys.argv[2])
M = int(sys.argv[3])
print(f'In file {filename} adding {M} empty rows starting from row {N}')

wb = openpyxl.load_workbook(filename)
sheet = wb[wb.sheetnames[0]]
first_data = {}
second_data = {}

for crow in range(sheet.max_row, N-1, -1):
    for ccolumn in range(1, sheet.max_column):
        sheet.cell(row=crow+M, column=ccolumn).value=sheet.cell(row=crow, column=ccolumn).value
for crow in range(N, N+M):
    for ccolumn in range(1, sheet.max_column):
        sheet.cell(row=crow, column=ccolumn).value=''

wb.save('edited_file.xlsx')
