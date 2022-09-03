#! python3
# the program should open a spreadsheet and write the cells of column A 
# into one text file, 
# the cells of column B into another text file, and so on
# usage in CLI: python3 spreadsheetTotxt.py "filename.xlsx"

import openpyxl, logging, sys
from openpyxl.utils import get_column_letter
logging.basicConfig(level=logging.DEBUG, format='''%(asctime)s -  %(levelname)s
-  %(message)s''')

filename = sys.argv[1]

wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']

for column in range(1, sheet.max_column+1):
	column_file = open(f'column_{get_column_letter(column)}.txt', 'w')
	for row in range(1, sheet.max_row+1):
		try:
			column_file.write(sheet.cell(row=row, column=column).value)
		except TypeError:
			logging.debug(f'''Type error in cell {get_column_letter(column)}{row} - 
			content is {str(sheet.cell(row=row, column=column).value)}''')
			continue
		column_file.write('\n')
	column_file.close()
